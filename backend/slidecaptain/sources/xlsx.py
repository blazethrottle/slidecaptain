"""XLSX 자료 추출 (설계서 3.1, 단계 5A 묶음 B 태스크 B1).

원본 엑셀 파일에서 값이 있는 셀만 골라 AI 프롬프트에 넣을 수 있는 UTF-8 텍스트로 바꾸는 순수 함수다.
파일 시스템을 건드리지 않는다: 원본 보존과 추출본 저장은 태스크 B2(업로드 라우트)가 맡는다.

- 값이 있는 셀만 "행 N: 열=값" 형태로 나열한다. 빈 셀까지 그리는 표 형식은 서식만 있는 넓은
  범위가 글자 수 상한을 잠식해 실제 데이터를 잘라낸다는 것이 1차 적대 리뷰에서 실측됐다
  (계획서 가정 3).
- 상한 5종(계획서 가정 4): 로드 전 검사 2종(압축 해제 총 크기, 시트 XML의 셀 요소 수)은 로드
  자체의 비용을 막기 위해 openpyxl을 부르기 전에 zipfile로 먼저 본다. 로드 후 검사 3종(보이는
  시트 수, 값 셀 수, 추출 글자 수)은 초과분을 "어디서 잘렸는지" 본문과 notes에 남기며 잘라낸다.
"""

from __future__ import annotations

import re
import zipfile
from datetime import date, datetime, time, timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

_HIDDEN_STATES = {"hidden", "veryHidden"}
_WORKSHEET_XML_PREFIX = "xl/worksheets/"

_CELL_ELEMENT_RE = re.compile(rb"<c[ />]")
_PERCENT_RE = re.compile(r"0(\.(0+))?%")
_LOCALE_MARKER_RE = re.compile(r"\[\$([^\]]*)\]")
_LOCALE_CODE_SUFFIX_RE = re.compile(r"-[0-9A-Za-z]+$")
_QUOTED_LITERAL_RE = re.compile(r'"([^"]*)"')
_NEWLINE_RE = re.compile(r"\r\n|\r|\n")
_CURRENCY_TOKENS = ("₩", "$", "€", "£", "¥", "원", "KRW", "USD", "EUR", "GBP", "JPY")


class XlsxLimits(BaseModel):
    """추출기가 강제하는 상한 5종(계획서 가정 4). 기본값은 마스터 플랜과 적대 리뷰 실측 근거."""

    max_declared_bytes: int = 50 * 1024 * 1024  # 로드 전: zip 항목 선언 크기 합
    max_cell_elements: int = 400_000  # 로드 전: 시트 XML의 <c 요소 수(서식뿐인 셀 포함)
    max_sheets: int = 30  # 로드 후: 보이는 시트 수
    max_value_cells: int = 200_000  # 로드 후: 값이 있는 셀 수(전체 시트 합계)
    max_chars: int = 60_000  # 로드 후: 추출 본문 글자 수(자료 전체 상한 100,000자 안에 여지를 남긴다)


DEFAULT_LIMITS = XlsxLimits()


class XlsxTooLarge(ValueError):
    """로드 전 검사(압축 해제 크기, 셀 요소 수)를 넘어 거절한 파일."""


class XlsxUnreadable(ValueError):
    """손상되었거나 openpyxl이 열지 못하는 파일(구버전 xls, xlsb 등도 여기로 온다)."""


class XlsxExtraction(BaseModel):
    """extract_xlsx의 결과. text는 AI 프롬프트에 그대로 넣을 UTF-8 문자열이다."""

    text: str
    sheets: int  # 추출본에 포함된 보이는 시트 수(시트 상한을 넘으면 잘린 뒤의 수)
    cells: int  # 추출본에 포함된 값 셀 수(값 셀 수나 글자 수 상한을 넘으면 잘린 뒤의 수)
    truncated: bool
    notes: list[str] = []


def extract_xlsx(data: bytes, filename: str, limits: XlsxLimits = DEFAULT_LIMITS) -> XlsxExtraction:
    """XLSX 바이트를 시트별 값 목록 텍스트로 바꾼다.

    파일이 상한을 넘으면 XlsxTooLarge, 손상되었거나 openpyxl이 열지 못하면 XlsxUnreadable을
    낸다(둘 다 ValueError 하위). 예외 재매핑은 화이트리스트가 아니다: XlsxTooLarge를 제외한
    모든 예외를 XlsxUnreadable로 바꾼다(계획서 B1. zipfile.BadZipFile, KeyError, OSError,
    xml.etree.ElementTree.ParseError 등 종류를 가리지 않는다).
    """
    try:
        _check_pre_load_limits(data, limits)
        wb_formulas = load_workbook(BytesIO(data), data_only=False)
        wb_values = load_workbook(BytesIO(data), data_only=True)
        return _build_extraction(wb_formulas, wb_values, filename, limits)
    except XlsxTooLarge:
        raise
    except Exception as e:
        raise XlsxUnreadable(
            f"엑셀 파일 {filename}을 읽지 못했습니다. 파일이 손상되었거나 지원하지 않는 형식입니다"
            "(구버전 xls, xlsb 등은 지원하지 않습니다). 엑셀에서 열어 xlsx 형식으로 다시 저장한 뒤 "
            "올려 주세요."
        ) from e


def _check_pre_load_limits(data: bytes, limits: XlsxLimits) -> None:
    """로드 전 검사 2종. 둘 다 openpyxl을 부르지 않고 zipfile만으로 판정해 큰 파일의 로드
    비용 자체를 막는다(계획서 가정 4: 서식만 있는 셀 100만 개짜리 파일이 로드에 4.5초와
    1.16GB를 쓴 실측에 대한 방어)."""
    with zipfile.ZipFile(BytesIO(data)) as zf:
        infos = zf.infolist()
        declared_total = sum(info.file_size for info in infos)
        if declared_total > limits.max_declared_bytes:
            raise XlsxTooLarge(
                f"엑셀 파일이 너무 큽니다(압축을 풀면 약 {declared_total / (1024 * 1024):,.1f}MB, "
                f"한도 {limits.max_declared_bytes / (1024 * 1024):,.0f}MB). 시트를 나누거나 "
                "불필요한 서식을 정리한 뒤 다시 올려 주세요."
            )

        cell_elements = 0
        for info in infos:
            if not info.filename.startswith(_WORKSHEET_XML_PREFIX) or not info.filename.endswith(".xml"):
                continue
            content = zf.read(info.filename)
            cell_elements += len(_CELL_ELEMENT_RE.findall(content))
            if cell_elements > limits.max_cell_elements:
                raise XlsxTooLarge(
                    f"엑셀 파일의 셀 개수(서식만 있는 셀 포함)가 너무 많습니다(한도 "
                    f"{limits.max_cell_elements:,}개). 빈 서식이 넓게 적용된 범위를 정리한 뒤 "
                    "다시 올려 주세요."
                )


def _build_extraction(wb_formulas, wb_values, filename: str, limits: XlsxLimits) -> XlsxExtraction:
    notes: list[str] = []
    all_sheets = list(wb_formulas.worksheets)
    hidden_names = [ws.title for ws in all_sheets if ws.sheet_state in _HIDDEN_STATES]
    visible_sheets = [ws for ws in all_sheets if ws.sheet_state not in _HIDDEN_STATES]

    truncated = False
    included_sheets = visible_sheets
    if len(visible_sheets) > limits.max_sheets:
        truncated = True
        notes.append(f"(한계: 시트 {limits.max_sheets + 1}개 이후 생략)")
        included_sheets = visible_sheets[: limits.max_sheets]

    if hidden_names:
        notes.append(f"숨김 시트 {len(hidden_names)}개 제외: " + ", ".join(hidden_names))

    body_lines: list[str] = []
    body_len = 0
    total_value_cells = 0
    calc_missing = 0
    newline_replacements = 0
    stopped = False
    processed_sheet_names: list[str] = []

    for ws in included_sheets:
        if stopped:
            break
        vws = wb_values[ws.title]
        rows, sheet_value_cells, value_range_desc, merges, sheet_calc_missing, sheet_newlines = _scan_sheet(
            ws, vws
        )
        calc_missing += sheet_calc_missing
        newline_replacements += sheet_newlines
        sheet_name = _clean_sheet_name(ws.title)
        processed_sheet_names.append(sheet_name)

        block_lines = [f"## 시트: {sheet_name} ({value_range_desc}, 값 셀 {sheet_value_cells}개)"]
        if merges:
            block_lines.append("병합: " + ", ".join(merges))
        for line in block_lines:
            body_lines.append(line)
            body_len += len(line) + 1

        for row_no, tokens in rows:
            row_cell_count = len(tokens)
            if total_value_cells + row_cell_count > limits.max_value_cells:
                truncated = True
                note = (
                    f"(한계: 값 셀 {limits.max_value_cells:,}개 초과분 생략. 시트 {sheet_name} 행 "
                    f"{row_no}부터)"
                )
                notes.append(note)
                body_lines.append(note)
                stopped = True
                break

            prefix = f"행 {row_no}: "
            full_line = prefix + " | ".join(tokens)
            if body_len + len(full_line) + 1 > limits.max_chars:
                truncated = True
                budget = limits.max_chars - body_len - len(prefix)
                kept, skipped_cols = _fit_tokens(tokens, budget)
                if kept:
                    body_lines.append(prefix + " | ".join(kept))
                    total_value_cells += len(kept)
                if skipped_cols:
                    notes.append(f"행 {row_no}에서 {', '.join(skipped_cols)}열 생략(글자 수 한도)")
                note = f"(한계: {limits.max_chars:,}자 초과분 생략. 시트 {sheet_name} 행 {row_no}부터)"
                notes.append(note)
                body_lines.append(note)
                stopped = True
                break

            body_lines.append(full_line)
            body_len += len(full_line) + 1
            total_value_cells += row_cell_count

    if stopped:
        remaining_sheets = included_sheets[len(processed_sheet_names) :]
        if remaining_sheets:
            remaining_names = [_clean_sheet_name(w.title) for w in remaining_sheets]
            notes.append("(한계: 이후 시트 생략: " + ", ".join(remaining_names) + ")")

    if calc_missing:
        notes.append(f"계산값 없음: {calc_missing}곳")
    if newline_replacements:
        notes.append(f"줄바꿈을 공백으로 바꿨습니다: {newline_replacements}곳")

    header_lines = [
        f"# XLSX 추출: {filename}",
        f"추출 시각: {datetime.now().astimezone().isoformat(timespec='seconds')}",
        f"시트 수: {len(processed_sheet_names)}개",
    ]
    if notes:
        header_lines.append("적용된 한계와 참고:")
        header_lines.extend(f"- {note}" for note in notes)

    text = "\n".join([*header_lines, "", *body_lines]) + "\n"

    return XlsxExtraction(
        text=text,
        sheets=len(processed_sheet_names),
        cells=total_value_cells,
        truncated=truncated,
        notes=notes,
    )


def _scan_sheet(ws, vws):
    """시트 하나를 훑어 값이 있는 셀만 행 단위로 모은다. 반환:
    (행 목록[(행 번호, [열=값 조각, ...])], 값 셀 수, "값 범위 A1:F120" 문구, 병합 범위 문구 목록,
    계산값 없음 건수, 줄바꿈 치환 건수). ws는 수식용(data_only=False), vws는 같은 시트의 값용
    (data_only=True) 워크시트다."""
    tokens_by_row: dict[int, list[str]] = {}
    min_row = min_col = max_row = max_col = None
    value_cell_count = 0
    calc_missing = 0
    newline_replacements = 0

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            r, c = cell.row, cell.column
            value_cell_count += 1
            min_row = r if min_row is None else min(min_row, r)
            max_row = r if max_row is None else max(max_row, r)
            min_col = c if min_col is None else min(min_col, c)
            max_col = c if max_col is None else max(max_col, c)

            if cell.data_type == "f":
                calc_value = vws.cell(row=r, column=c).value
                if calc_value is None:
                    calc_missing += 1
                    calc_text = "(계산값 없음)"
                else:
                    calc_text, nl = _format_value(calc_value, cell.number_format)
                    newline_replacements += nl
                formula_text, formula_nl = _format_string(cell.value)
                newline_replacements += formula_nl
                value_text = f"{formula_text} → {calc_text}"
            elif cell.hyperlink is not None:
                display, nl = _format_value(cell.value, cell.number_format)
                newline_replacements += nl
                target = cell.hyperlink.target or cell.hyperlink.location or ""
                value_text = f"[{display}]({target})"
            else:
                value_text, nl = _format_value(cell.value, cell.number_format)
                newline_replacements += nl

            tokens_by_row.setdefault(r, []).append(f"{get_column_letter(c)}={value_text}")

    if value_cell_count == 0:
        value_range_desc = "값 범위 없음"
    else:
        value_range_desc = f"값 범위 {get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"

    merges = [str(rng) for rng in ws.merged_cells.ranges]
    rows = [(r, tokens_by_row[r]) for r in sorted(tokens_by_row)]
    return rows, value_cell_count, value_range_desc, merges, calc_missing, newline_replacements


def _clean_sheet_name(name: str) -> str:
    return name.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")


def _fit_tokens(tokens: list[str], budget: int) -> tuple[list[str], list[str]]:
    """budget 글자 안에 들어가는 만큼만 " | "로 이어질 완전한 토큰을 앞에서부터 채우고,
    나머지는 열 문자만 뽑아 돌려준다(한 토큰을 중간에서 끊지 않는다. 계획서 가정 4)."""
    kept: list[str] = []
    used = 0
    for i, token in enumerate(tokens):
        addition = token if not kept else " | " + token
        if used + len(addition) > budget:
            return kept, [t.split("=", 1)[0] for t in tokens[i:]]
        kept.append(token)
        used += len(addition)
    return kept, []


def _format_value(value, number_format: str) -> tuple[str, int]:
    """cell.value의 파이썬 타입과 number_format으로 표시 문자열을 만든다. (문자열, 줄바꿈
    치환 건수)를 돌려준다. 순서: 불리언(정수의 서브클래스라 정수보다 먼저) → 날짜시각 →
    날짜 → 시각 → 경과시간 → 문자열 → 정수, 실수(계획서 가정 3)."""
    if isinstance(value, bool):
        return ("참" if value else "거짓"), 0
    if isinstance(value, datetime):
        if _has_time_marker(number_format):
            return value.strftime("%Y-%m-%dT%H:%M:%S"), 0
        return value.strftime("%Y-%m-%d"), 0
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d"), 0
    if isinstance(value, time):
        return value.strftime("%H:%M:%S"), 0
    if isinstance(value, timedelta):
        return _format_timedelta(value), 0
    if isinstance(value, str):
        return _format_string(value)
    if isinstance(value, (int, float)):
        return _format_numeric(value, number_format), 0
    return str(value), 0


def _format_string(value: str) -> tuple[str, int]:
    """백슬래시를 먼저, 파이프를 다음에 이스케이프하고, 줄바꿈은 공백으로 바꾸며 건수를
    센다(계획서 가정 3 표기 순서)."""
    text = value.replace("\\", "\\\\").replace("|", "\\|")
    newline_count = len(_NEWLINE_RE.findall(text))
    text = _NEWLINE_RE.sub(" ", text)
    return text, newline_count


def _format_numeric(value: int | float, number_format: str) -> str:
    percent_match = _PERCENT_RE.search(number_format)
    if percent_match:
        decimals = len(percent_match.group(2)) if percent_match.group(2) else 0
        return f"{value} ({value * 100:.{decimals}f}%)"
    currency = _currency_symbol(number_format)
    if currency:
        return f"{value} ({currency})"
    return str(value)


def _has_time_marker(number_format: str) -> bool:
    lowered = number_format.lower()
    return "h" in lowered or "s" in lowered or ":" in lowered


def _format_timedelta(value: timedelta) -> str:
    total_seconds = int(value.total_seconds())
    sign = "-" if total_seconds < 0 else ""
    total_seconds = abs(total_seconds)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"


def _currency_symbol(number_format: str) -> str | None:
    """서식 문자열에서 통화 기호를 찾는다. [$...] 로케일 마커가 있으면 그 안(끝의 -로케일코드는
    제거)에서만 찾고, 마커 밖의 $ 는 보지 않는다(마커 자체가 $ 로 시작하는 문법이라 오탐이
    난다. 계획서 가정 3). 마커가 없으면 따옴표 리터럴과 서식 문자열 전체에서 찾는다."""
    marker = _LOCALE_MARKER_RE.search(number_format)
    if marker is not None:
        inner = _LOCALE_CODE_SUFFIX_RE.sub("", marker.group(1)).strip()
        for token in _CURRENCY_TOKENS:
            if token in inner:
                return token
        return _quoted_currency(number_format)
    quoted = _quoted_currency(number_format)
    if quoted:
        return quoted
    for token in _CURRENCY_TOKENS:
        if token in number_format:
            return token
    return None


def _quoted_currency(number_format: str) -> str | None:
    for literal in _QUOTED_LITERAL_RE.findall(number_format):
        for token in _CURRENCY_TOKENS:
            if token in literal:
                return token
    return None
