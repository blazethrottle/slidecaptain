"""XLSX 추출기 테스트 (태스크 B1, 계획서 "테스트" 절 전항목).

픽스처는 저장소에 바이너리를 두지 않고 이 파일과 xlsx_fixtures.py가 그때그때 만든다.
"""

import datetime

import openpyxl
import pytest
from openpyxl.styles import PatternFill

import xlsx_fixtures
from slidecaptain.sources.xlsx import (
    XlsxLimits,
    XlsxTooLarge,
    XlsxUnreadable,
    extract_xlsx,
)


# -- 주 시나리오: 문자열, 정수, 실수, 날짜류, 수식, 링크, 병합이 섞인 시트 --------------------


def _primary_fixture_bytes() -> bytes:
    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws.title = "매출"
        ws["A1"] = "항목"
        ws["B1"] = "금액"
        ws["A2"] = "식비"
        ws["B2"] = 15000
        ws["A3"] = "교통비"
        ws["B3"] = 3.5
        ws["B4"] = "비고"
        ws.merge_cells("B4:C4")
        ws["A5"] = "날짜만"
        ws["B5"] = datetime.date(2026, 9, 4)
        ws["C5"] = datetime.datetime(2026, 9, 4, 9, 5, 0)
        ws["D5"] = datetime.time(9, 5, 0)
        ws["E5"] = datetime.timedelta(hours=30, minutes=5)
        ws["A6"] = "자세히"
        ws["A6"].hyperlink = "https://example.com/detail"
        ws["B9"] = "=SUM(B2:B3)"
        wb.create_sheet("메모")["A1"] = "부가 설명"

    return xlsx_fixtures.workbook_bytes(build)


def test_primary_fixture_extracts_all_value_types_without_blank_rows():
    result = extract_xlsx(_primary_fixture_bytes(), "매출현황.xlsx")

    assert result.sheets == 2
    assert "## 시트: 매출 (값 범위 A1:E9, 값 셀 14개)" in result.text
    assert "병합: B4:C4" in result.text
    # 문자열과 정수, 실수가 그대로 나온다
    assert "행 1: A=항목 | B=금액" in result.text
    assert "행 2: A=식비 | B=15000" in result.text
    assert "행 3: A=교통비 | B=3.5" in result.text
    assert "행 4: B=비고" in result.text
    # 날짜만 / 날짜+시각 / 시각만 / 경과시간
    assert "B=2026-09-04" in result.text
    assert "C=2026-09-04T09:05:00" in result.text
    assert "D=09:05:00" in result.text
    assert "E=30:05:00" in result.text
    # 하이퍼링크 표기
    assert "A=[자세히](https://example.com/detail)" in result.text
    # 수식은 열 문자만 토큰 키로 쓴다(행 번호는 "행 9:"가 이미 가리킨다). 이 픽스처(일반
    # openpyxl 저장)에는 캐시된 계산값이 없다
    assert "행 9: B==SUM(B2:B3) → (계산값 없음)" in result.text
    # 두 번째 시트
    assert "## 시트: 메모 (값 범위 A1:A1, 값 셀 1개)" in result.text
    assert "행 1: A=부가 설명" in result.text
    # 빈 행과 빈 셀은 나오지 않는다 (값 없는 행 7, 8은 등장하지 않는다)
    assert "행 7:" not in result.text
    assert "행 8:" not in result.text
    assert result.truncated is False


def test_primary_fixture_header_has_filename_and_timestamp():
    result = extract_xlsx(_primary_fixture_bytes(), "매출현황.xlsx")
    assert result.text.startswith("# XLSX 추출: 매출현황.xlsx\n")
    assert "추출 시각: " in result.text
    assert "시트 수: 2개" in result.text


def test_float_that_looks_like_integer_is_preserved_via_xml_patch():
    """3.0은 openpyxl이 저장할 때 <v>3</v>로 정규화하므로(계획서 가정 3), XML을 사후
    치환해 원본에 소수점이 있는 경우를 재현한다."""

    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws["C1"] = 3.0

    data = xlsx_fixtures.workbook_bytes(build)
    patched = xlsx_fixtures.replace_in_member(
        data, "xl/worksheets/sheet1.xml", ('<c r="C1" t="n"><v>3</v></c>', '<c r="C1" t="n"><v>3.0</v></c>')
    )
    result = extract_xlsx(patched, "정수형실수.xlsx")
    assert "C=3.0" in result.text


# -- 서식만 있는 넓은 범위 (1차 리뷰 critical 재현) ------------------------------------


def test_value_range_ignores_formatting_only_cells():
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        for row in range(1, 991):
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = fill
        ws.cell(row=995, column=1, value="결과값")

    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "서식만.xlsx")

    assert "값 범위 A995:A995, 값 셀 1개" in result.text
    assert "행 995: A=결과값" in result.text
    # 서식만 있는 990행이 표에 반영되지 않으므로 추출본이 짧다
    assert len(result.text) < 2000
    assert result.truncated is False


# -- 계산값 캐시 -----------------------------------------------------------------------


def test_missing_calc_cache_reports_placeholder_and_note():
    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws["B2"] = 1
        ws["B3"] = 2
        ws["B9"] = "=SUM(B2:B3)"

    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "계산없음.xlsx")
    assert "(계산값 없음)" in result.text
    assert any("계산값 없음: 1곳" in n for n in result.notes)


def test_calc_cache_value_shown_when_xml_patched():
    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws["B2"] = 1
        ws["B3"] = 2
        ws["B9"] = "=SUM(B2:B3)"

    data = xlsx_fixtures.workbook_bytes(build)
    patched = xlsx_fixtures.replace_in_member(
        data, "xl/worksheets/sheet1.xml", ('<c r="B9"><f>SUM(B2:B3)</f><v></v></c>', '<c r="B9"><f>SUM(B2:B3)</f><v>3</v></c>')
    )
    result = extract_xlsx(patched, "계산있음.xlsx")
    assert "B==SUM(B2:B3) → 3" in result.text
    assert not any("계산값 없음" in n for n in result.notes)


def test_formula_source_text_is_escaped_like_string_values():
    """수식 원문에 파이프 문자가 있으면(TEXTJOIN 등 흔한 패턴) 문자열 값과 같은 이스케이프
    규칙을 적용한다 (B1 리뷰 formula-text-unescaped-pipe: 계산값과 하이퍼링크 표시 텍스트는
    이스케이프되는데 수식 원문만 그대로 삽입돼 " | " 토큰 구분자와 구별되지 않았다)."""

    def build(wb: openpyxl.Workbook) -> None:
        wb.active["A1"] = '=TEXTJOIN("|",TRUE,A1:A5)'

    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "수식이스케이프.xlsx")
    assert 'A==TEXTJOIN("\\|",TRUE,A1:A5) → (계산값 없음)' in result.text


# -- 숨김 시트 -------------------------------------------------------------------------


def test_hidden_and_very_hidden_sheets_excluded_and_named_in_header():
    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws.title = "매출"
        ws["A1"] = "값"
        hidden = wb.create_sheet("내부검토")
        hidden.sheet_state = "hidden"
        very_hidden = wb.create_sheet("구버전")
        very_hidden.sheet_state = "veryHidden"

    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "숨김포함.xlsx")

    assert result.sheets == 1
    assert "숨김 시트 2개 제외: 내부검토, 구버전" in result.text
    assert any("숨김 시트 2개 제외: 내부검토, 구버전" in n for n in result.notes)
    # 숨김 시트는 이름만 머리에 적힐 뿐, 자기 시트 구획은 본문에 없다
    assert "## 시트: 내부검토" not in result.text
    assert "## 시트: 구버전" not in result.text


def test_sheet_name_newline_becomes_space():
    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws.title = "메모\n시트"
        ws["A1"] = "값"

    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "개행시트.xlsx")
    assert "## 시트: 메모 시트 (" in result.text
    assert "메모\n시트" not in result.text


# -- 값 표기: 이스케이프, 줄바꿈, 퍼센트, 통화, 불리언 ----------------------------------


def test_pipe_and_backslash_are_escaped_in_order():
    # 백슬래시를 먼저, 파이프를 다음에 이스케이프한다 (계획서 가정 3)
    raw = "가격" + "|" + "단위" + "\\" + "개수"
    expected = "가격" + "\\|" + "단위" + "\\\\" + "개수"

    def build(wb: openpyxl.Workbook) -> None:
        wb.active["A1"] = raw

    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "이스케이프.xlsx")
    assert f"A={expected}" in result.text


def test_newline_in_string_is_replaced_and_counted():
    def build(wb: openpyxl.Workbook) -> None:
        wb.active["A1"] = "첫줄\n둘째줄"

    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "줄바꿈.xlsx")
    assert "A=첫줄 둘째줄" in result.text
    assert any("줄바꿈을 공백으로 바꿨습니다: 1곳" in n for n in result.notes)


@pytest.mark.parametrize(
    ("number_format", "value", "expected"),
    [
        ("0.0%", 0.153, "0.153 (15.3%)"),
        ("0.00%", 0.1234, "0.1234 (12.34%)"),
    ],
)
def test_percent_format_decimals_follow_pattern(number_format, value, expected):
    def build(wb: openpyxl.Workbook) -> None:
        cell = wb.active["A1"]
        cell.value = value
        cell.number_format = number_format

    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "퍼센트.xlsx")
    assert f"A={expected}" in result.text


@pytest.mark.parametrize(
    ("number_format", "value", "expected"),
    [
        ("[$₩-412]#,##0", 1234, "1234 (₩)"),
        ('#,##0"원"', 5000, "5000 (원)"),
        ("[$EUR ]#,##0.00", 10.5, "10.5 (EUR)"),
    ],
)
def test_currency_format_appends_symbol(number_format, value, expected):
    def build(wb: openpyxl.Workbook) -> None:
        cell = wb.active["A1"]
        cell.value = value
        cell.number_format = number_format

    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "통화.xlsx")
    assert f"A={expected}" in result.text


def test_locale_marker_dollar_syntax_is_not_a_false_positive():
    """[$₩-412]의 $ 는 마커 문법이지 통화 기호가 아니므로 ₩ 만 잡히고 $ 는 함께 붙지 않는다."""

    def build(wb: openpyxl.Workbook) -> None:
        cell = wb.active["A1"]
        cell.value = 100
        cell.number_format = "[$₩-412]#,##0"

    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "원화.xlsx")
    assert "A=100 (₩)" in result.text
    assert "($)" not in result.text


def test_boolean_shown_as_korean_words():
    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws["A1"] = True
        ws["A2"] = False

    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "불리언.xlsx")
    assert "A=참" in result.text
    assert "A=거짓" in result.text


# -- 상한 5종 ---------------------------------------------------------------------------


def test_sheet_cap_truncates_after_default_thirty():
    def build(wb: openpyxl.Workbook) -> None:
        wb.active.title = "s0"
        wb.active["A1"] = "첫시트"
        for i in range(1, 31):
            wb.create_sheet(f"s{i}")["A1"] = f"값{i}"

    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "시트31개.xlsx")

    assert result.sheets == 30
    assert result.truncated is True
    assert "(한계: 시트 31개 이후 생략)" in result.text
    assert any("시트 31개 이후 생략" in n for n in result.notes)
    assert "## 시트: s0 " in result.text
    assert "## 시트: s30 " not in result.text  # 31번째(0-indexed 30)는 생략된다


def test_value_cell_cap_truncates_and_notes_location():
    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        for row in range(1, 21):
            ws.cell(row=row, column=1, value=row)

    limits = XlsxLimits(max_value_cells=10)
    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "값셀상한.xlsx", limits)

    assert result.truncated is True
    assert result.cells == 10
    assert "행 11:" not in result.text
    assert "값 셀 10개 초과분 생략" in result.text
    assert "행 11부터" in result.text


def test_char_cap_truncates_mid_row_keeping_whole_tokens():
    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        for col in range(1, 201):
            ws.cell(row=1, column=col, value=col)

    limits = XlsxLimits(max_value_cells=10_000, max_chars=120)
    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "글자수상한.xlsx", limits)

    assert result.truncated is True
    assert "A=1" in result.text  # 앞쪽 열은 살아남는다
    assert "GR=200" not in result.text  # 마지막 열은 잘린다
    assert any("열 생략(글자 수 한도)" in n for n in result.notes)
    assert any("자 초과분 생략" in n and "행 1부터" in n for n in result.notes)
    # 살아남은 셀 수가 cells 필드에 반영된다 (B1 리뷰 cells-undercount-charcap: 이 경로가
    # total_value_cells를 갱신하지 않고 바로 break해 0을 돌려주고 있었다)
    assert result.cells > 0


def test_char_cap_mid_sheet_drops_later_sheet_from_count_and_notes():
    """글자 수 상한이 첫 시트 도중에 걸리면, 아직 처리되지 않은 이후 시트는 시트 수 집계에서
    빠지고 그 이름이 notes에 남는다 (B1 리뷰 sheets-silent-drop: 이전에는 sheets 필드와
    "시트 수" 머리글이 트렁케이션 전 개수를 그대로 보고하고 생략 사실이 어디에도 남지 않았다)."""

    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        ws.title = "s1"
        for col in range(1, 51):
            ws.cell(row=1, column=col, value=col)
        wb.create_sheet("s2")["A1"] = "다른시트값"

    limits = XlsxLimits(max_value_cells=10_000, max_chars=100)
    result = extract_xlsx(xlsx_fixtures.workbook_bytes(build), "다중시트글자수.xlsx", limits)

    assert result.truncated is True
    assert result.sheets == 1
    assert "시트 수: 1개" in result.text
    assert "다른시트값" not in result.text
    assert "## 시트: s2" not in result.text
    assert any("이후 시트 생략" in n and "s2" in n for n in result.notes)


def test_declared_size_cap_rejects_before_load():
    data = xlsx_fixtures.workbook_bytes(lambda wb: wb.active.__setitem__("A1", "x"))
    patched = xlsx_fixtures.patch_declared_size(data, "xl/workbook.xml", 2_000_000)
    limits = XlsxLimits(max_declared_bytes=1_000)

    with pytest.raises(XlsxTooLarge):
        extract_xlsx(patched, "너무큼.xlsx", limits)


def test_cell_element_cap_rejects_formatting_only_sheet_before_load():
    fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")

    def build(wb: openpyxl.Workbook) -> None:
        ws = wb.active
        for row in range(1, 21):
            for col in range(1, 11):
                ws.cell(row=row, column=col).fill = fill

    limits = XlsxLimits(max_cell_elements=50)
    with pytest.raises(XlsxTooLarge):
        extract_xlsx(xlsx_fixtures.workbook_bytes(build), "셀요소상한.xlsx", limits)


# -- 손상된 입력 -------------------------------------------------------------------------


def test_random_bytes_is_unreadable_with_korean_message():
    with pytest.raises(XlsxUnreadable) as exc:
        extract_xlsx(xlsx_fixtures.not_a_zip_bytes(), "깨짐.xlsx")
    assert "엑셀 파일" in str(exc.value)


def test_valid_zip_without_xlsx_structure_is_unreadable():
    with pytest.raises(XlsxUnreadable):
        extract_xlsx(xlsx_fixtures.zip_without_content_types(), "zip아님.xlsx")


def test_crc_mismatch_is_unreadable():
    data = xlsx_fixtures.workbook_bytes(lambda wb: wb.active.__setitem__("A1", "x"))
    corrupted = xlsx_fixtures.corrupt_crc(data, "xl/worksheets/sheet1.xml")
    with pytest.raises(XlsxUnreadable):
        extract_xlsx(corrupted, "crc깨짐.xlsx")


def test_xlsb_labeled_as_xlsx_is_unreadable():
    with pytest.raises(XlsxUnreadable):
        extract_xlsx(xlsx_fixtures.pseudo_xlsb_bytes(), "가짜.xlsx")


def test_empty_workbook_has_only_sheet_title():
    result = extract_xlsx(xlsx_fixtures.workbook_bytes(lambda wb: None), "빈파일.xlsx")
    assert result.sheets == 1
    assert result.cells == 0
    assert "값 범위 없음, 값 셀 0개" in result.text
    assert "행 " not in result.text  # 값 있는 행이 없으니 행 단위 줄 자체가 없다
    assert result.notes == []
